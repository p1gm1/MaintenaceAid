# Pytest
import pytest
import unittest

# Django
from django.conf import settings

from openpyxl import load_workbook

pytestmark = pytest.mark.django_db

class TestVibrationsExcelForm:

    def test_excel_operations(self):
        excel_file = load_workbook(str(settings.APPS_DIR) + '/static/exs.xlsx')

        assertions = unittest.TestCase('__init__')

        assertions.assertIsNotNone(excel_file)

        worksheet = excel_file['Sheet1']

        mp_arr = []
        i = 0

        for row in worksheet.iter_rows():
            if i != 0:
                for i in range(len(row)):
                    assertions.assertIsNotNone(row[i].value)
                    if i == 0:
                        mp_arr.append({'monitoring point': row[i].value})
                        assertions.assertEqual(type(row[i].value),type('s'))
                    if i == 1:
                        mp_arr.append({f'p{i}': row[i].value})
                        assertions.assertEqual(type(row[i].value),type(1.00))
            i += 1