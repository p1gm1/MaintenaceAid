# Pyxl
from openpyxl import load_workbook

def excel_to_dict(path_to_file):
    """Handles data to excel
    from database.
    """
    excel_file=load_workbook(path_to_file)
    worksheet=excel_file['Sheet1']

    mp_arr = []
    i=0

    for row in worksheet.iter_rows():
        if i != 0:
            for j in range(len(row)):
                if j == 0:
                    monitoring_point=str(row[j].value).strip()
                elif j == 1:
                    velocity=float(row[j].value)
                elif j == 2:
                    aceleration=float(row[j].value)
                elif j == 3:
                    dem_odulada=float(row[j].value)
                
            mp_arr.append({"monitoring_point": monitoring_point,
                           "velocity": velocity,
                           "aceleration": aceleration,
                           "dem_odulada": dem_odulada})
        else:
            i += 1
    print(mp_arr)
    return mp_arr